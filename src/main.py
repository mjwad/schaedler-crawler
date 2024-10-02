"""This module defines the main entry point for the Apify Actor.

Feel free to modify this file to suit your specific needs.

To build Apify Actors, utilize the Apify SDK toolkit, read more at the official documentation:
https://docs.apify.com/sdk/python
"""
import asyncio
from urllib.parse import urljoin, urlparse
from playwright.async_api import async_playwright
from apify import Actor, Request
import pandas as pd
from openpyxl import load_workbook

# Initialize the Excel file path and other configurations
EXCEL_FILE_PATH = './import_template.xlsx'  # Path to the existing Excel file

# Load the workbook and the specific sheet you want to modify
workbook = load_workbook(EXCEL_FILE_PATH)
sheet = workbook.active

# Read the existing data using pandas
df_existing = pd.read_excel(EXCEL_FILE_PATH)

# Get the headers from the first row of the Excel sheet
headers = {sheet.cell(row=2, column=i).value: i for i in range(1, sheet.max_column + 1)}

# Helper function to scrape product details
async def scrape_product_details(page, product_url):
    try:
        Actor.log.info(f'Scraping product details from: {product_url}')
        await page.wait_for_selector('main', timeout=10000)

        # Select the main product container
        product_container = await page.query_selector('main')

        # Example product data extraction (adjust selectors accordingly)
        product_img_selector = await product_container.query_selector('img[data-test-selector="productDetails_mainImage"]')
        product_img_url = await product_img_selector.get_attribute('src') if product_img_selector else None
        
        # product name selector
        product_name_select = await product_container.query_selector('h2[data-test-selector^="productDetails_productDescription"]')
        product_name = await product_name_select.inner_text() if product_name_select else None
        
        # manufactor article number
        mfg_value_selector = await product_container.query_selector('p:has(span:text("MFG #"))')
        mfg_value = (await mfg_value_selector.inner_text()).replace('MFG #', '').strip() if mfg_value_selector else None
        
        # sku selector
        sku_selector = await product_container.query_selector('p:has(span:text("SKU #"))')
        sku_value = (await sku_selector.inner_text()).replace('SKU #', '').strip() if sku_selector else None
        
        # upc selector
        upc_selector = await product_container.query_selector('p:has(span:text("UPC #"))')
        upc_value = (await upc_selector.inner_text()).replace('UPC #', '').strip() if upc_selector else None
        
        # production description
        product_description_selector = await product_container.query_selector('div[data-test-selector="productDetails_htmlContent"]')
        product_desciption = await product_description_selector.inner_text() if product_description_selector else None


        # Extract attributes section
        header_section_selector = 'dt[data-test-selector="sectionHeader"]:has-text("Attributes")'
        await page.wait_for_selector(f'{header_section_selector} > button', timeout=10000)
        attributes_btn_selector = await product_container.query_selector(f'{header_section_selector} > button')

        if attributes_btn_selector:
            await attributes_btn_selector.click()
            section_panel = await product_container.query_selector(f'{header_section_selector} ~ dd[data-test-selector="sectionPanel"]')

            # Extract manufacturer and original data type
            manufacturer_selector =  await section_panel.query_selector('div:has(span:has-text("Brand")) > span:nth-child(2)') 
            if manufacturer_selector:
                manufacturer = await manufacturer_selector.inner_text()

            original_data_2 = await section_panel.query_selector('div:has(span:text-is("Type: ")) > span:nth-child(2)')
            if original_data_2:
                original_data_2 = await original_data_2.inner_text()

        return {
            "product_img_url": product_img_url,
            "product_name": product_name,
            "mfg_value": mfg_value,
            "sku": sku_value,
            "upc": upc_value,
            "product_description": product_desciption,
            "original_data_2": original_data_2,
            "manufacturer": manufacturer
        }

    except Exception as e:
        Actor.log.error(f"Error scraping product details from {product_url}: {e}")
        return None

async def scrape_product_urls(page):
    """Scrapes product URLs from the product listing page."""
    await page.wait_for_selector('[data-test-selector*="productListProductCard"]')
    
    page_parsed_url = urlparse(page.url)

    # Select all product cards and filter out unwanted ones
    products = [
        product for product in await page.query_selector_all('div[data-test-selector^="productListProductCard"]')
        if not await product.evaluate('p => p.closest("[data-test-selector=\'cardListProductsNarrow\']")')
    ]

    product_urls = []
    for product in products:
        product_url_selector = await product.query_selector('a')
        relative_url_path = await product_url_selector.get_attribute('href')
        product_url = urljoin(f"{page_parsed_url.scheme}://{page_parsed_url.netloc}", relative_url_path)
        product_urls.append(product_url)

    return product_urls

# Function to find the last non-empty row
def find_last_non_empty_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value is not None for cell in sheet[row]):
            return row
    return 0  # In case the sheet is completely empty

# scrapping all products on a page
async def main() -> None:
    """Main entry point for the Apify Actor."""
    async with Actor:
        actor_input = await Actor.get_input() or {}
        start_urls = actor_input.get('start_urls', [{'url': 'https://apify.com'}])
        max_depth = actor_input.get('max_depth', 1)

        request_queue = await Actor.open_request_queue()

        for start_url in start_urls:
            url = start_url.get('url')
            await request_queue.add_request(Request.from_url(url, user_data={'depth': 0, 'product_link_page': True}))
        
        # List to store all product data
        all_products_data = []

        Actor.log.info('Launching Playwright...')
        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=Actor.config.headless)
            context = await browser.new_context()


            # Fetch each request from the request queue
            while request := await request_queue.fetch_next_request():
                url = request.url
                depth = request.user_data['depth']
                product_link_page = request.user_data['product_link_page']
                
                page = await context.new_page()
                await page.goto(url)

                try:
                    if product_link_page:
                        product_urls = await scrape_product_urls(page)

                        # Scrape product URLs and add to request queue
                        for url in product_urls:
                            await request_queue.add_request(Request.from_url(url, user_data={
                                'depth': depth, 
                                'product_link_page': False
                            }))
                    else:                            
                        # Scrape product details
                        product_data = await scrape_product_details(page, url)

                        # Map product data to Excel columns
                        mapped_product_data = {
                            "Originaldata 1": product_data["upc"],
                            "Originaldata 2": product_data["original_data_2"],
                            "Manufacturer": product_data["manufacturer"],
                            "Manufacturer ArticleNumber": product_data["mfg_value"],
                            "ProductName": product_data["product_name"],
                            "ProductDescription": product_data["product_description"],
                            "ProductImageLink": product_data["product_img_url"],
                            "Supplier ArticleNumber": product_data["sku"],
                        }

                        # Add the mapped product data to the list
                        all_products_data.append(mapped_product_data)

                except Exception as e:
                    Actor.log.error(f"Error processing page {url}: {e}")
                finally:
                    await page.close()
            await browser.close()

            # After the loop finishes, save all the collected product data to the Excel file

        try:
            # Load the existing Excel file
            df = pd.read_excel(EXCEL_FILE_PATH)
            Actor.log.info(f'Existing Excel file loaded. {df}')
        except FileNotFoundError:
            # If file not found, create a new DataFrame
            df = pd.DataFrame(columns=[
                "ProductImageLink", 
                "ProductName", 
                "Manufacturer ArticleNumber", 
                "Supplier ArticleNumber", 
                "Originaldata 1", 
                "ProductDescription", 
                "Manufacturer", 
                "Originaldata 2"
            ])
            Actor.log.info('Excel file not found. A new file will be created.')

        # Create a pandas DataFrame for the new product data
        df_new = pd.DataFrame(all_products_data)

        # Find the last non-empty row
        last_row = find_last_non_empty_row(sheet)

        # Append the new data row by row into the correct columns based on the headers
        for index, row in df_new.iterrows():
            for column_name, value in row.items():
                if column_name in headers:  # Check if the column exists in the headers
                    col_index = headers[column_name]  # Get the correct column index
                    sheet.cell(row=last_row + index + 1, column=col_index, value=value)

        # Save the file without changing its styling
        workbook.save(EXCEL_FILE_PATH)

        # Save the file without changing its styling
        workbook.save(EXCEL_FILE_PATH)
        print(f"Data successfully appended to {EXCEL_FILE_PATH}")

        # Gracefully exit the actor
        await Actor.exit()